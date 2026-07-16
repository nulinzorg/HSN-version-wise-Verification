"""
hsn_core.py
===========
Shared logic used by both app.py (interactive dashboard) and
headless_check.py (scheduled/automated run). Keeping this in one place
means the dashboard and the automated email always compare data the
exact same way.
"""

import re
import unicodedata
from io import BytesIO
from datetime import datetime

import pandas as pd
import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

SOURCE_URL = "https://tutorial.gst.gov.in/downloads/HSN_SAC.xlsx"

# The GST portal (like many government sites) blocks requests that don't
# look like they're coming from a real browser - Python's requests
# library sends a very identifiable default header, which some sites
# reject outright with a 403. Sending headers that mimic an ordinary
# browser avoids that.
DOWNLOAD_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream,*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://services.gst.gov.in/services/searchhsnsac",
}

INVISIBLE_CHARS_PATTERN = re.compile(
    r"["
    r"\u200b\u200c\u200d\u2060\ufeff"   # zero-width space/joiner/BOM
    r"\xa0"                              # non-breaking space
    r"\u2000-\u200a"                     # various unicode spaces
    r"\u202f\u205f\u3000"                # narrow/medium/ideographic space
    r"\x00-\x08\x0b\x0c\x0e-\x1f"        # control characters
    r"]"
)
XML_RESERVED_PATTERN = re.compile(r'[<>&"\']')

FILL_ADDED = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_DELETED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_MODIFIED = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def fetch_source_excel(url: str = SOURCE_URL) -> BytesIO:
    last_error = None
    for attempt in range(2):  # one retry - some sites transiently 403/503 on bot-detection heuristics
        try:
            resp = requests.get(url, headers=DOWNLOAD_HEADERS, timeout=30)
            resp.raise_for_status()
            content_type = resp.headers.get("Content-Type", "")
            if "html" in content_type.lower():
                # The site returned an HTML page (e.g. an error/login/captcha
                # page) instead of the actual spreadsheet - this looks like
                # success (status 200) but isn't a real Excel file.
                raise ValueError(
                    "The GST portal returned an HTML page instead of the Excel file "
                    "(likely a block/redirect page). Try opening the link in a browser "
                    "to confirm it still works, or the portal may have changed its URL."
                )
            return BytesIO(resp.content)
        except (requests.exceptions.RequestException, ValueError) as e:
            last_error = e
            continue
    raise RuntimeError(
        f"Could not download the HSN/SAC file from the GST portal after 2 attempts: {last_error}"
    )


def load_excel(file_like) -> pd.DataFrame:
    """Reads a single-sheet Excel file (or just the first sheet if there
    happen to be more) as a flat DataFrame. Used for reference files and
    for re-loading a single already-separated sheet snapshot."""
    return pd.read_excel(file_like, dtype=str).fillna("")


def load_excel_sheets(file_like) -> dict:
    """Reads EVERY sheet in the workbook and returns them as separate
    DataFrames, keyed by sheet name - sheets are NEVER merged or
    cross-compared with each other. This matters because the portal's
    HSN_SAC.xlsx may have HSN and SAC entries on distinct sheets, and
    they represent genuinely different code systems that should each be
    tracked (and diffed against their own history) independently."""
    return {name: df.fillna("") for name, df in pd.read_excel(file_like, sheet_name=None, dtype=str).items()}


def normalize_for_compare(val) -> str:
    if val is None:
        return ""
    s = str(val)
    s = unicodedata.normalize("NFKC", s)
    s = INVISIBLE_CHARS_PATTERN.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def compare_dataframes(ref_df, src_df, key_col, compare_cols):
    ref = ref_df.copy()
    src = src_df.copy()
    ref["_key"] = ref[key_col].apply(normalize_for_compare)
    src["_key"] = src[key_col].apply(normalize_for_compare)

    merged = ref.merge(
        src, on="_key", how="outer", suffixes=("_ref", "_src"), indicator=True
    )

    added = merged[merged["_merge"] == "right_only"].copy()
    deleted = merged[merged["_merge"] == "left_only"].copy()
    both = merged[merged["_merge"] == "both"].copy()

    modified_idx = []
    changed_map = {}
    for idx, row in both.iterrows():
        changed_fields = []
        for col in compare_cols:
            rcol, scol = f"{col}_ref", f"{col}_src"
            if rcol in row and scol in row:
                if normalize_for_compare(row[rcol]) != normalize_for_compare(row[scol]):
                    changed_fields.append(col)
        if changed_fields:
            modified_idx.append(idx)
            changed_map[idx] = ", ".join(changed_fields)

    modified = both.loc[modified_idx].copy()
    modified["_changed_fields"] = modified.index.map(changed_map)

    return added, deleted, modified


def detect_invisible_issues(df: pd.DataFrame, columns, source_label: str) -> pd.DataFrame:
    issues = []
    for idx, row in df.iterrows():
        for col in columns:
            if col not in row:
                continue
            raw = row[col]
            if raw is None or raw == "":
                continue
            s = str(raw)
            problems = []
            if INVISIBLE_CHARS_PATTERN.search(s):
                problems.append("invisible/unicode-space character")
            if s != s.strip():
                problems.append("leading/trailing whitespace")
            if XML_RESERVED_PATTERN.search(s):
                problems.append("XML-reserved character (&, <, >, \", ')")
            if problems:
                issues.append(
                    {
                        "source": source_label,
                        "row_index": idx,
                        "column": col,
                        "value_repr": repr(s),
                        "issues": "; ".join(problems),
                    }
                )
    return pd.DataFrame(issues)


FILL_ISSUE = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


def export_validation_highlighted_excel(df: pd.DataFrame, key_col: str, check_cols) -> bytes:
    """Returns an Excel workbook containing ONLY the rows that have at
    least one invisible/special-character issue in check_cols, with the
    exact offending cell(s) highlighted in red - not just a flat list of
    values, but the actual data in context so it's immediately clear
    which row and which field needs fixing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Description Issues"

    problem_rows = []
    for idx, row in df.iterrows():
        row_issues = {}
        for col in check_cols:
            if col not in df.columns:
                continue
            raw = row[col]
            if raw is None or raw == "":
                continue
            s = str(raw)
            problems = []
            if INVISIBLE_CHARS_PATTERN.search(s):
                problems.append("invisible character")
            if s != s.strip():
                problems.append("leading/trailing whitespace")
            if XML_RESERVED_PATTERN.search(s):
                problems.append("special character")
            if problems:
                row_issues[col] = "; ".join(problems)
        if row_issues:
            problem_rows.append((row, row_issues))

    header = [key_col] + list(check_cols) + ["Issues Found"]
    ws.append(header)

    if not problem_rows:
        ws.append(["No invisible or special characters found in any row."] + [""] * len(check_cols) + [""])
    else:
        for row, row_issues in problem_rows:
            row_vals = [row.get(key_col, "")] + [row.get(c, "") for c in check_cols]
            issues_summary = "; ".join(f"{c}: {msg}" for c, msg in row_issues.items())
            ws.append(row_vals + [issues_summary])
            excel_row_num = ws.max_row
            for col_offset, col_name in enumerate(check_cols, start=2):  # column 1 = key_col
                if col_name in row_issues:
                    ws.cell(row=excel_row_num, column=col_offset).fill = FILL_ISSUE

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(max_len + 2, 10), 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def clean_value(val) -> str:
    s = str(val)
    s = unicodedata.normalize("NFKC", s)
    s = INVISIBLE_CHARS_PATTERN.sub("", s)
    return s.strip()


def build_xml(df: pd.DataFrame, export_cols, row_tag="HSNItem", root_tag="HSN_SAC_Delta", auto_clean=False):
    root = etree.Element(root_tag)
    root.set("generated_on", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    for _, row in df.iterrows():
        item = etree.SubElement(root, row_tag)
        for col in export_cols:
            if col not in row:
                continue
            tag_name = re.sub(r"\W+", "_", col).strip("_") or "field"
            child = etree.SubElement(item, tag_name)
            val = row[col]
            val = "" if val is None else str(val)
            child.text = clean_value(val) if auto_clean else val
    return etree.tostring(root, pretty_print=True, xml_declaration=True, encoding="UTF-8")


def export_highlighted_excel(added, deleted, modified, display_cols_map):
    wb = Workbook()
    wb.remove(wb.active)

    def write_sheet(name, df, fill, cols):
        ws = wb.create_sheet(name)
        if df.empty:
            ws.append(["No changes"])
            return
        clean_df = df[cols].copy()
        for r in dataframe_to_rows(clean_df, index=False, header=True):
            ws.append(r)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.fill = fill

    write_sheet("Added", added, FILL_ADDED, display_cols_map["added"])
    write_sheet("Deleted", deleted, FILL_DELETED, display_cols_map["deleted"])
    write_sheet("Modified", modified, FILL_MODIFIED, display_cols_map["modified"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_delta_for_xml(added, modified, key_src_col):
    """Combines Added + Modified rows (source-side values only) into one
    frame ready for XML export - deleted/unchanged rows are never included."""
    added_cols = [c for c in added.columns if c.endswith("_src") or c == key_src_col]
    modified_src_cols = [c for c in modified.columns if c.endswith("_src")]
    if key_src_col not in modified_src_cols and key_src_col in modified.columns:
        modified_src_cols = modified_src_cols + [key_src_col]

    delta = pd.concat(
        [
            added[added_cols].rename(columns=lambda c: c.replace("_src", "")),
            modified[modified_src_cols].rename(columns=lambda c: c.replace("_src", "")),
        ],
        ignore_index=True,
    )
    return delta
